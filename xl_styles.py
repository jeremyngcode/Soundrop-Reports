from openpyxl.styles import Font, PatternFill
# -------------------------------------------------------------------------------------------------

boldunderline = Font(
	bold=True,
	underline='single'
)
biu = Font(
	bold=True,
	italic=True,
	underline='single'
)

grey_fill = PatternFill(
	patternType='solid',
	fgColor='808080'
)
orange_fill = PatternFill(
	patternType='solid',
	fgColor='FFC000'
)
yellow_fill = PatternFill(
	patternType='solid',
	fgColor='FFFF00'
)
